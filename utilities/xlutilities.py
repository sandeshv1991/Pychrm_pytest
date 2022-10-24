import openpyxl
from openpyxl import workbook
from openpyxl.styles import PatternFill


def getrowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)


def getcolumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)


def read_data(file, sheetname, rownum, colnum):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook[sheetname]
    return sheet.cell(rownum, colnum).value


def write_data(file, sheetname, rownum, colnum, data):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook['Data']
    sheet.cell(rownum, colnum).value = data
    Workbook.save(file)


def fillGreencolour(file, sheetname, rownum, colnum):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook['Data']
    greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rownum, colnum).fill = greenfill
    Workbook.save(file)


def fillRedcolour(file, sheetname, rownum, colnum):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook['Data']
    redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, colnum).fill = redfill
    Workbook.save(file)
